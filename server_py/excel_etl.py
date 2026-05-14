"""
Smart Workflows — AI-translated mini-ETL for Excel files.

Two-phase design:

Phase 1 (setup, AI involved):
    POST /etl/translate
        body: { prompt, slot_filenames }, files: <multiple uploads>
        → Gemini parses the prompt + sees the sample file schemas
        → returns a validated workflow JSON
        → also returns the input_schemas snapshot

Phase 2 (daily run, NO AI):
    POST /etl/run-init
        body: { workflow_json, input_schemas, slot_files: [...] }
        → uploads daily files, validates them against input_schemas
        → returns rename suggestions for any drifted columns
        → caches the resolved DataFrames in-memory under a run_id

    POST /etl/run-step
        body: { run_id, step_index, column_remap }
        → executes steps[0..step_index] on the cached frames
        → returns the top-N preview rows + match-rate stats

    POST /etl/run-finalize
        body: { run_id }
        → executes the full pipeline, returns the xlsx as base64

Operations whitelist (v1):  read_excel (implicit), join, select, rename, fill, write_excel.
"""

from __future__ import annotations

import io
import json
import os
import re
import time
import uuid
import base64
import difflib
from threading import Lock as _Lock
from typing import Any, Dict, List, Optional, Literal

import pandas as pd
from fastapi import APIRouter, File, Form, HTTPException, UploadFile
from pydantic import BaseModel, Field, ValidationError

try:
    from google import genai
    from google.genai import types as genai_types
    _GENAI_AVAILABLE = True
except Exception:
    _GENAI_AVAILABLE = False


router = APIRouter(prefix="/etl", tags=["Smart Workflows"])


# -----------------------------------------------------------
# Constants & cache
# -----------------------------------------------------------

MAX_FILE_BYTES = 100 * 1024 * 1024
MAX_ROWS = 500_000
PREVIEW_ROWS = 10
SAMPLE_ROWS_FOR_AI = 5
CACHE_TTL_SEC = 60 * 60  # 1 hour

# Run cache: run_id → { frames: { slotName: df }, workflow: {...}, expires_at, derived: { stepName: df } }
_RUN_CACHE: Dict[str, Dict[str, Any]] = {}
_CACHE_LOCK = _Lock()


def _cache_set(run_id: str, payload: Dict[str, Any]) -> None:
    payload["expires_at"] = time.time() + CACHE_TTL_SEC
    with _CACHE_LOCK:
        # Sweep expired entries
        now = time.time()
        for k in list(_RUN_CACHE.keys()):
            if _RUN_CACHE[k]["expires_at"] < now:
                del _RUN_CACHE[k]
        _RUN_CACHE[run_id] = payload


def _cache_get(run_id: str) -> Optional[Dict[str, Any]]:
    with _CACHE_LOCK:
        entry = _RUN_CACHE.get(run_id)
        if entry is None or entry["expires_at"] < time.time():
            return None
        return entry


# -----------------------------------------------------------
# Workflow schema (Pydantic — validates AI output)
# -----------------------------------------------------------

class InputSlot(BaseModel):
    slot: str                          # short identifier — "prices", "master", "vans"
    filename: Optional[str] = ""        # remembered original filename for hint on daily run
    sheet: Optional[str] = ""           # which Excel sheet (default first)
    description: Optional[str] = ""     # human label — "Daily price list"


# Each step references prior step outputs by name (the `as` field) or input slots.
# Only the operations below are accepted — anything else is rejected.

class JoinStep(BaseModel):
    op: Literal["join"]
    left: str                          # slot name or step "as" name
    right: str
    on_left: str                       # column in `left`
    on_right: str                      # column in `right`
    how: Literal["left", "inner"] = "left"
    as_: str = Field(alias="as")        # name to reference this step's output downstream

    class Config:
        populate_by_name = True


class FillStep(BaseModel):
    op: Literal["fill"]
    input: str                          # slot or "as" name
    column: str
    value: Any                          # can be number, string, or null-marker
    as_: str = Field(alias="as")

    class Config:
        populate_by_name = True


class RenameStep(BaseModel):
    op: Literal["rename"]
    input: str
    map: Dict[str, str]                 # { "Old": "New" }
    as_: str = Field(alias="as")

    class Config:
        populate_by_name = True


class SelectStep(BaseModel):
    op: Literal["select"]
    input: str
    columns: List[str]                  # output column order
    as_: str = Field(alias="as")

    class Config:
        populate_by_name = True


# Discriminated union — Pydantic v1/v2 compat handled by trying each shape.
WORKFLOW_OP_TYPES = {"join": JoinStep, "fill": FillStep, "rename": RenameStep, "select": SelectStep}


class WorkflowOutput(BaseModel):
    final_step: str                     # which step output to write
    columns: Optional[List[str]] = None  # optional column projection
    filename: str = "result.xlsx"
    sheet: str = "Result"
    # Output shape — what the user wants back:
    #   "records"  → raw rows xlsx (default, original behaviour)
    #   "value"    → single scalar (KPI) computed from `agg_column` with `agg_func`
    #   "template" → records xlsx with branded header + auto column widths
    output_type: Literal["records", "value", "template"] = "records"
    agg_func: Optional[Literal["sum", "count", "avg", "min", "max"]] = None
    agg_column: Optional[str] = None    # required when output_type == "value"
    template_title: Optional[str] = None  # heading printed in template mode


class Workflow(BaseModel):
    inputs: List[InputSlot]
    steps: List[Dict[str, Any]]         # parsed/validated separately because of discriminated union
    output: WorkflowOutput


def _validate_steps(steps: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """Validate each step against its op-specific schema. Returns sanitized list of dicts."""
    out: List[Dict[str, Any]] = []
    for i, raw in enumerate(steps):
        op = raw.get("op")
        cls = WORKFLOW_OP_TYPES.get(op)
        if cls is None:
            raise HTTPException(400, f"Step {i}: unknown op '{op}'. Allowed: {list(WORKFLOW_OP_TYPES)}")
        try:
            parsed = cls(**raw)
        except ValidationError as e:
            raise HTTPException(400, f"Step {i} ({op}): {e}")
        # Re-emit as dict with `as` key (not `as_`)
        d = parsed.model_dump(by_alias=True)
        out.append(d)
    return out


# -----------------------------------------------------------
# File parsing
# -----------------------------------------------------------

def _read_pdf_tables(content: bytes, filename: str) -> pd.DataFrame:
    """Extract tabular data from a PDF using pdfplumber.

    Strategy: walk every page, collect every table. Use the first table's first row as the
    canonical header. Append rows from any subsequent table whose column count matches.
    This handles the common case of multi-page invoices / statements / reports where the
    same table layout repeats across pages.
    """
    try:
        import pdfplumber
    except ImportError:
        raise HTTPException(500, "PDF support not installed. Run: pip install pdfplumber")

    header: Optional[List[str]] = None
    rows: List[List[Any]] = []
    skipped_tables = 0

    try:
        with pdfplumber.open(io.BytesIO(content)) as pdf:
            for page in pdf.pages:
                page_tables = page.extract_tables() or []
                for tbl in page_tables:
                    if not tbl or len(tbl) < 1:
                        continue
                    if header is None:
                        header = [str(c).strip() if c is not None else f"col_{i}" for i, c in enumerate(tbl[0])]
                        # If header has duplicates, suffix them
                        seen: Dict[str, int] = {}
                        deduped: List[str] = []
                        for h in header:
                            if h in seen:
                                seen[h] += 1
                                deduped.append(f"{h}_{seen[h]}")
                            else:
                                seen[h] = 1
                                deduped.append(h)
                        header = deduped
                        rows.extend([row for row in tbl[1:] if any(c is not None and str(c).strip() for c in row)])
                    else:
                        if len(tbl[0]) == len(header):
                            # Same shape — assume it's a continuation. If the first row looks
                            # like a repeated header, drop it.
                            first = [str(c or "").strip() for c in tbl[0]]
                            data_start = 1 if first == header else 0
                            rows.extend([row for row in tbl[data_start:] if any(c is not None and str(c).strip() for c in row)])
                        else:
                            skipped_tables += 1
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(400, f"Failed to parse PDF {filename}: {e}")

    if header is None or not rows:
        raise HTTPException(400, f"{filename}: no extractable tables found in PDF. PDFs without table grids (e.g. scanned images) need OCR first.")

    df = pd.DataFrame(rows, columns=header)
    # Strip whitespace from string cells
    for c in df.columns:
        if df[c].dtype == object:
            df[c] = df[c].map(lambda v: str(v).strip() if v is not None else v)
    return df


def _read_excel_or_csv(content: bytes, filename: str, sheet: Optional[str] = None) -> pd.DataFrame:
    name = (filename or "").lower()
    try:
        if name.endswith(".csv"):
            try:
                df = pd.read_csv(io.BytesIO(content))
            except UnicodeDecodeError:
                df = pd.read_csv(io.BytesIO(content), encoding="latin-1")
        elif name.endswith((".xlsx", ".xls")):
            df = pd.read_excel(io.BytesIO(content), sheet_name=sheet or 0)
        elif name.endswith(".pdf"):
            df = _read_pdf_tables(content, filename)
        else:
            raise HTTPException(415, f"Unsupported file type: {filename}. Use .xlsx, .xls, .csv, or .pdf")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(400, f"Failed to parse {filename}: {e}")
    if len(df) > MAX_ROWS:
        raise HTTPException(413, f"{filename}: row count {len(df):,} exceeds {MAX_ROWS:,}")
    # Trim whitespace column headers
    df.columns = [str(c).strip() for c in df.columns]
    return df


def _schema_snapshot(df: pd.DataFrame, filename: str = "", sheet: str = "") -> Dict[str, Any]:
    """Lightweight schema fingerprint used to detect drift on daily runs."""
    cols: List[Dict[str, Any]] = []
    for c in df.columns:
        s = df[c]
        sample = [str(v)[:40] for v in s.dropna().head(3).tolist()]
        cols.append({"name": str(c), "dtype": str(s.dtype), "sample": sample})
    return {"filename": filename, "sheet": sheet, "columns": cols, "row_count": int(len(df))}


def _df_records_safe(df: pd.DataFrame) -> List[Dict[str, Any]]:
    """JSON-safe records (NaN → null, datetimes → iso)."""
    sanitized = df.where(pd.notnull(df), None)
    out: List[Dict[str, Any]] = []
    for row in sanitized.to_dict(orient="records"):
        clean = {}
        for k, v in row.items():
            if isinstance(v, float) and (v != v or v in (float("inf"), float("-inf"))):
                clean[str(k)] = None
            elif hasattr(v, "isoformat"):
                clean[str(k)] = v.isoformat()
            else:
                clean[str(k)] = v
        out.append(clean)
    return out


# -----------------------------------------------------------
# Gemini — translate the prompt → workflow JSON
# -----------------------------------------------------------

def _gemini_client() -> Optional[Any]:
    key = os.getenv("GEMINI_API_KEY") or os.getenv("GOOGLE_API_KEY")
    if not key or not _GENAI_AVAILABLE:
        return None
    try:
        return genai.Client(api_key=key)
    except Exception as e:
        print(f"[excel_etl] gemini init failed: {e}")
        return None


def _is_quota_error(msg: str) -> bool:
    return "RESOURCE_EXHAUSTED" in msg or "429" in msg or "quota" in msg.lower()


def _quota_or_502(e: Exception, label: str) -> HTTPException:
    """Translate a Gemini SDK exception into a clean HTTPException.

    The free-tier 429 from Google returns BOTH per-minute and per-day quota violations
    in a single message. We detect which is the binding constraint and produce an
    honest message — "retry in 59s" is misleading when the daily cap is the real blocker
    (waiting 59s and retrying just hits the daily quota again, looping forever).
    """
    msg = str(e)
    if not _is_quota_error(msg):
        return HTTPException(502, f"{label}: {msg[:300]}")

    # Detect the binding quota. The "PerDay" / "Daily" quota beats per-minute when both fire,
    # because waiting out the per-minute window won't unblock the per-day one.
    daily_hit = bool(re.search(r"PerDay|Daily", msg, re.IGNORECASE))
    # `limit: 0` in Google's response means the free tier is effectively unavailable
    # for this model on this project — common for newly added models or some regions.
    free_tier_zero = bool(re.search(r"limit:\s*0\b", msg))

    if daily_hit or free_tier_zero:
        return HTTPException(
            429,
            "Gemini free-tier daily quota exhausted. "
            "Waiting will not help today — either enable billing on your "
            "Google AI Studio key (https://aistudio.google.com/) or wait for "
            "the daily reset (~midnight Pacific).",
        )

    # Per-minute only — short retry helps
    retry_match = re.search(r"retryDelay['\":\s]+(\d+)s", msg)
    retry_hint = f" Retry in ~{retry_match.group(1)}s." if retry_match else " Retry in a moment."
    return HTTPException(429, f"Gemini API rate-limited (per-minute).{retry_hint}")


_TRANSLATE_SYSTEM = """You are a Data Engineer that translates plain-English instructions into a structured ETL workflow JSON.

You may ONLY emit operations from this allowed list:
  - "join":    merge two tables on a key.   Fields: op, left, right, on_left, on_right, how (left|inner), as.
  - "rename":  rename columns.              Fields: op, input, map (old→new dict), as.
  - "fill":    replace nulls in a column.   Fields: op, input, column, value, as.
  - "select":  keep only certain columns.   Fields: op, input, columns (list), as.

Each step's "as" gives that step's output a name. Later steps reference it via "input" / "left" / "right".
Inputs may also reference an `inputs[].slot` name directly.

You will be given:
  1. A user prompt describing the task in plain language.
  2. A list of input file slots. Each slot has a `slot` name and a column list with sample values.

Your job: emit a JSON document matching this exact shape:

{
  "inputs": [
    { "slot": "<short_id>", "filename": "<original filename>", "sheet": "<sheet or empty>", "description": "<human label>" },
    ...
  ],
  "steps": [
    { "op": "join",   "left": "<slot|as>", "right": "<slot|as>", "on_left": "<col>", "on_right": "<col>", "how": "left", "as": "step1" },
    { "op": "fill",   "input": "<slot|as>", "column": "<col>", "value": 0, "as": "step2" },
    { "op": "select", "input": "<slot|as>", "columns": ["A","B","C","D"], "as": "final" }
  ],
  "output": {
    "final_step": "<as name of last step>",
    "columns": ["A","B","C","D"],
    "filename": "<reasonable_default>.xlsx",
    "sheet": "Result"
  }
}

Rules (CRITICAL):
- Use exact column names from the provided slot schemas. Never invent columns.
- For each join, choose the on_left and on_right keys that share the same identity (e.g. an Item Code on both sides).
- The final step's columns should match the user's requested output columns.
- ALWAYS chain steps via the "as" field. The last step's "as" must equal output.final_step.
- Return ONLY JSON. No markdown fences, no commentary.
"""


def _ai_translate(
    prompt: str,
    slot_schemas: List[Dict[str, Any]],
    relationships: Optional[List[Dict[str, Any]]] = None,
) -> Optional[Dict[str, Any]]:
    client = _gemini_client()
    if client is None:
        return None
    rel_block = ""
    if relationships:
        rel_block = (
            f"CONFIRMED RELATIONSHIPS (use these as join keys when joining the listed slot pairs):\n"
            f"{json.dumps(relationships, default=str)[:2000]}\n\n"
        )
    body = (
        f"USER PROMPT:\n{prompt}\n\n"
        f"INPUT FILE SLOTS:\n{json.dumps(slot_schemas, default=str)[:8000]}\n\n"
        f"{rel_block}"
        "Emit the workflow JSON now."
    )
    try:
        resp = client.models.generate_content(
            model="gemini-2.0-flash",
            contents=[_TRANSLATE_SYSTEM, body],
            config=genai_types.GenerateContentConfig(
                response_mime_type="application/json",
                temperature=0.1,
                max_output_tokens=2500,
            ),
        )
        raw = (resp.text or "").strip()
        if raw.startswith("```"):
            raw = re.sub(r"^```[a-zA-Z]*\n?", "", raw)
            raw = re.sub(r"\n?```$", "", raw)
        return json.loads(raw)
    except Exception as e:
        msg = str(e)
        print(f"[excel_etl] translate failed: {msg[:300]}")
        # Re-raise quota errors as a clean HTTP 429 so the client doesn't see a generic 503
        if _is_quota_error(msg):
            raise _quota_or_502(e, "Translate failed")
        return None


# -----------------------------------------------------------
# Workflow execution (deterministic, no AI)
# -----------------------------------------------------------

class StepStats(BaseModel):
    matched_rows: int = 0
    unmatched_rows: int = 0
    total_rows: int = 0
    note: str = ""


def _exec_step(
    step: Dict[str, Any],
    frames: Dict[str, pd.DataFrame],
) -> tuple[pd.DataFrame, StepStats]:
    op = step["op"]
    stats = StepStats()

    def _get(name: str) -> pd.DataFrame:
        if name not in frames:
            raise HTTPException(400, f"Step references unknown frame '{name}'")
        return frames[name]

    if op == "join":
        left = _get(step["left"]).copy()
        right = _get(step["right"]).copy()
        on_left = step["on_left"]
        on_right = step["on_right"]
        how = step.get("how", "left")
        if on_left not in left.columns:
            raise HTTPException(400, f"Join: left column '{on_left}' not in '{step['left']}'")
        if on_right not in right.columns:
            raise HTTPException(400, f"Join: right column '{on_right}' not in '{step['right']}'")
        # Coerce both keys to string so '123' (int) and '123' (str) match.
        left["__k"] = left[on_left].astype(str)
        right["__k"] = right[on_right].astype(str)
        # Drop the rename target if the right key has the same name as a left column (avoid suffix mess).
        if on_right != on_left and on_right in left.columns:
            right = right.drop(columns=[on_right], errors="ignore")
        merged = left.merge(right.drop(columns=["__k"]).assign(__k=right["__k"]), on="__k", how=how, suffixes=("", "_r"))
        merged = merged.drop(columns=["__k"], errors="ignore")
        # Stats
        right_keys = set(right["__k"].dropna().tolist())
        left_keys = left["__k"].dropna()
        matched = int(left_keys.isin(right_keys).sum()) if len(left_keys) else 0
        stats.total_rows = int(len(left))
        stats.matched_rows = matched
        stats.unmatched_rows = stats.total_rows - matched
        return merged, stats

    if op == "fill":
        df = _get(step["input"]).copy()
        col = step["column"]
        val = step["value"]
        if col not in df.columns:
            stats.note = f"column '{col}' not present — fill skipped"
            return df, stats
        before = int(df[col].isna().sum())
        df[col] = df[col].fillna(val)
        stats.total_rows = int(len(df))
        stats.note = f"filled {before} null values in '{col}' with {val!r}"
        return df, stats

    if op == "rename":
        df = _get(step["input"]).copy()
        m = step.get("map", {}) or {}
        present = {k: v for k, v in m.items() if k in df.columns}
        df = df.rename(columns=present)
        stats.total_rows = int(len(df))
        stats.note = f"renamed {len(present)} of {len(m)} columns"
        return df, stats

    if op == "select":
        df = _get(step["input"]).copy()
        cols = step.get("columns", []) or []
        missing = [c for c in cols if c not in df.columns]
        present = [c for c in cols if c in df.columns]
        df = df[present].copy()
        stats.total_rows = int(len(df))
        if missing:
            stats.note = f"dropped {len(missing)} missing columns: {missing[:5]}"
        return df, stats

    raise HTTPException(400, f"Unsupported op '{op}'")


def _execute(
    workflow: Dict[str, Any],
    inputs: Dict[str, pd.DataFrame],
    up_to: Optional[int] = None,
) -> Dict[str, Any]:
    """Run steps 0..up_to (inclusive). Returns dict with frames per step + stats."""
    frames: Dict[str, pd.DataFrame] = dict(inputs)
    steps = workflow.get("steps", [])
    cap = len(steps) - 1 if up_to is None else min(up_to, len(steps) - 1)
    step_results: List[Dict[str, Any]] = []
    for i, s in enumerate(steps):
        if i > cap:
            break
        df, stats = _exec_step(s, frames)
        as_name = s.get("as") or f"step{i}"
        frames[as_name] = df
        step_results.append({
            "index": i,
            "op": s["op"],
            "as": as_name,
            "stats": stats.model_dump(),
            "preview_columns": [str(c) for c in df.columns],
            "preview_rows": _df_records_safe(df.head(PREVIEW_ROWS)),
            "row_count": int(len(df)),
        })
    return {"frames": frames, "step_results": step_results}


# -----------------------------------------------------------
# Endpoints
# -----------------------------------------------------------

@router.post("/translate")
async def translate(
    prompt: str = Form(...),
    slot_names: str = Form(...),  # JSON list, parallel to files
    files: List[UploadFile] = File(...),
    relationships: str = Form("[]"),  # JSON list of confirmed joins (optional)
):
    """Phase 1: take a natural-language prompt + sample files → AI workflow JSON."""
    if not prompt.strip():
        raise HTTPException(400, "Prompt is required")
    try:
        slot_list = json.loads(slot_names)
    except Exception:
        raise HTTPException(400, "slot_names must be a JSON list")
    if not isinstance(slot_list, list) or len(slot_list) != len(files):
        raise HTTPException(400, f"slot_names ({len(slot_list)}) must match files ({len(files)})")

    # Parse each file + capture schema snapshot
    slot_schemas: List[Dict[str, Any]] = []
    input_schemas: Dict[str, Dict[str, Any]] = {}
    sample_for_ai: List[Dict[str, Any]] = []
    for slot, file in zip(slot_list, files):
        content = await file.read()
        if len(content) > MAX_FILE_BYTES:
            raise HTTPException(413, f"{file.filename}: too large")
        df = _read_excel_or_csv(content, file.filename or slot)
        snap = _schema_snapshot(df, filename=file.filename or "", sheet="")
        input_schemas[slot] = snap
        # AI sees: slot name, filename, columns + a tiny sample
        sample_for_ai.append({
            "slot": slot,
            "filename": file.filename,
            "columns": [{"name": c["name"], "dtype": c["dtype"], "sample": c["sample"]} for c in snap["columns"]],
            "sample_rows": _df_records_safe(df.head(SAMPLE_ROWS_FOR_AI)),
        })
        slot_schemas.append(snap)

    try:
        rels = json.loads(relationships) if relationships else []
        if not isinstance(rels, list):
            rels = []
    except Exception:
        rels = []

    workflow_raw = _ai_translate(prompt, sample_for_ai, rels)
    if workflow_raw is None:
        raise HTTPException(503, "AI translation unavailable. Is GEMINI_API_KEY set?")

    # Validate
    try:
        steps_clean = _validate_steps(workflow_raw.get("steps", []))
        output = workflow_raw.get("output") or {}
        # Provide sane defaults if AI omitted output fields
        output.setdefault("final_step", steps_clean[-1].get("as") if steps_clean else "")
        output.setdefault("filename", "result.xlsx")
        output.setdefault("sheet", "Result")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(400, f"AI returned invalid workflow: {e}")

    workflow_json = {
        "inputs": [
            {"slot": s, "filename": input_schemas[s]["filename"], "sheet": "", "description": ""}
            for s in slot_list
        ],
        "steps": steps_clean,
        "output": output,
    }

    return {
        "workflow_json": workflow_json,
        "input_schemas": input_schemas,
    }


@router.post("/detect-relationships")
async def detect_relationships(
    slot_names: str = Form(...),
    files: List[UploadFile] = File(...),
):
    """Suggest VLookup-style joins between uploaded files.

    Heuristic: for every pair (slot A, slot B), every column pair (a, b):
      - score = name_similarity * 0.5 + value_overlap * 0.5
      - keep only pairs with score >= 0.55 and value_overlap >= 0.30
    Returns the top suggestions per file pair so the UI can draw the diagram.
    """
    try:
        slots = json.loads(slot_names)
    except Exception:
        raise HTTPException(400, "slot_names must be a JSON list")
    if not isinstance(slots, list) or len(slots) != len(files):
        raise HTTPException(400, "slot_names must match files")

    # Parse all files once
    parsed: Dict[str, Dict[str, Any]] = {}
    for slot, file in zip(slots, files):
        content = await file.read()
        if len(content) > MAX_FILE_BYTES:
            raise HTTPException(413, f"{file.filename}: too large")
        df = _read_excel_or_csv(content, file.filename or slot)
        # Build per-column value-set (string-coerced, capped) for overlap calc
        col_sets: Dict[str, set] = {}
        for c in df.columns:
            vals = df[c].dropna().astype(str).head(2000).tolist()
            col_sets[str(c)] = set(vals)
        parsed[slot] = {
            "filename": file.filename,
            "columns": [str(c) for c in df.columns],
            "col_sets": col_sets,
            "row_count": int(len(df)),
        }

    relationships: List[Dict[str, Any]] = []
    slot_list = list(parsed.keys())
    for i in range(len(slot_list)):
        for j in range(i + 1, len(slot_list)):
            a, b = slot_list[i], slot_list[j]
            best: List[Dict[str, Any]] = []
            for ca in parsed[a]["columns"]:
                for cb in parsed[b]["columns"]:
                    sa: set = parsed[a]["col_sets"][ca]
                    sb: set = parsed[b]["col_sets"][cb]
                    if not sa or not sb:
                        continue
                    overlap = len(sa & sb) / max(1, min(len(sa), len(sb)))
                    name_sim = difflib.SequenceMatcher(None, ca.lower(), cb.lower()).ratio()
                    score = name_sim * 0.5 + overlap * 0.5
                    if score >= 0.55 and overlap >= 0.30:
                        best.append({
                            "left_column": ca,
                            "right_column": cb,
                            "overlap": round(overlap, 3),
                            "name_similarity": round(name_sim, 3),
                            "score": round(score, 3),
                        })
            best.sort(key=lambda r: r["score"], reverse=True)
            if best:
                relationships.append({
                    "left_slot": a,
                    "right_slot": b,
                    "suggestions": best[:3],   # top 3 candidate keys per pair
                })

    return {
        "files": [
            {"slot": s, "filename": parsed[s]["filename"], "columns": parsed[s]["columns"], "row_count": parsed[s]["row_count"]}
            for s in slot_list
        ],
        "relationships": relationships,
    }


_ENHANCE_SYSTEM = """You rewrite a user's rough natural-language data task into a precise, professional ETL brief.

Input:
  - The user's draft prompt (often informal, may have typos).
  - The list of uploaded file slots (slot name, filename, columns).
  - Optionally, confirmed key relationships between the files (slot.column ↔ slot.column).

Your output is a SINGLE rewritten prompt — readable English (or Arabic if input is Arabic), structured but conversational.

It must:
  1. Reference each file by its slot name in backticks, e.g. `prices`, `master`.
  2. State explicit join keys when relationships are given. Format: "Join `prices` with `master` on `Item Code`."
  3. List the desired final columns clearly when the draft mentions them.
  4. Spell out null-handling, filtering, and renames as discrete sentences.
  5. Be 3–8 short sentences. No headings, no bullets, no markdown.

Return ONLY the rewritten prompt text. No prefix, no quotes, no commentary.
"""


@router.post("/enhance-prompt")
async def enhance_prompt(body: Dict[str, Any]):
    """Rewrite a rough draft prompt into a polished pro version using the file context."""
    draft: str = (body.get("prompt") or "").strip()
    slots: List[Dict[str, Any]] = body.get("slots") or []
    relationships: List[Dict[str, Any]] = body.get("relationships") or []
    if not draft:
        raise HTTPException(400, "prompt is required")

    client = _gemini_client()
    if client is None:
        raise HTTPException(503, "AI unavailable. Is GEMINI_API_KEY set?")

    body_text = (
        f"DRAFT PROMPT:\n{draft}\n\n"
        f"FILE SLOTS:\n{json.dumps(slots, default=str)[:4000]}\n\n"
        f"CONFIRMED RELATIONSHIPS:\n{json.dumps(relationships, default=str)[:2000]}\n\n"
        "Rewrite the draft into the polished form now."
    )
    try:
        resp = client.models.generate_content(
            model="gemini-2.0-flash",
            contents=[_ENHANCE_SYSTEM, body_text],
            config=genai_types.GenerateContentConfig(temperature=0.2, max_output_tokens=600),
        )
        text = (resp.text or "").strip()
        # Strip any accidental markdown fences / surrounding quotes
        text = re.sub(r"^```[a-zA-Z]*\n?", "", text)
        text = re.sub(r"\n?```$", "", text)
        text = text.strip().strip('"').strip("'")
        if not text:
            raise HTTPException(502, "AI returned an empty rewrite")
        return {"enhanced_prompt": text}
    except HTTPException:
        raise
    except Exception as e:
        raise _quota_or_502(e, "Enhance failed")


@router.post("/run-init")
async def run_init(
    workflow_json: str = Form(...),
    input_schemas: str = Form(...),
    slot_names: str = Form(...),  # JSON list, parallel to files
    files: List[UploadFile] = File(...),
):
    """Phase 2 step 1: upload daily files → validate vs setup-time schemas → return run_id + drift report."""
    try:
        wf = json.loads(workflow_json)
        original_schemas = json.loads(input_schemas)
        slots = json.loads(slot_names)
    except Exception as e:
        raise HTTPException(400, f"Bad form payload: {e}")

    if len(slots) != len(files):
        raise HTTPException(400, "slot_names length must match files length")

    frames: Dict[str, pd.DataFrame] = {}
    drift: Dict[str, Any] = {}
    for slot, file in zip(slots, files):
        content = await file.read()
        if len(content) > MAX_FILE_BYTES:
            raise HTTPException(413, f"{file.filename}: too large")
        df = _read_excel_or_csv(content, file.filename or slot)
        frames[slot] = df

        original = original_schemas.get(slot, {})
        original_cols = [c["name"] for c in original.get("columns", [])]
        new_cols = [str(c) for c in df.columns]
        # Fuzzy-suggest renames for missing original columns
        missing: List[Dict[str, Any]] = []
        for orig in original_cols:
            if orig in new_cols:
                continue
            # Best fuzzy match among the new columns
            matches = difflib.get_close_matches(orig, new_cols, n=3, cutoff=0.4)
            missing.append({"original": orig, "suggestions": matches})
        drift[slot] = {
            "filename": file.filename,
            "original_columns": original_cols,
            "current_columns": new_cols,
            "missing_columns": missing,  # original-side columns not present in today's file
            "extra_columns": [c for c in new_cols if c not in original_cols],
            "row_count": int(len(df)),
        }

    run_id = uuid.uuid4().hex
    _cache_set(run_id, {"frames": frames, "workflow": wf})

    return {
        "run_id": run_id,
        "drift": drift,
    }


@router.post("/run-step")
async def run_step(body: Dict[str, Any]):
    """Phase 2 step N: execute steps[0..step_index] and return preview rows + stats."""
    run_id = body.get("run_id", "")
    step_index = int(body.get("step_index", 0))
    column_remap = body.get("column_remap") or {}  # { slot: { oldName: newName } }

    entry = _cache_get(run_id)
    if not entry:
        raise HTTPException(410, "Run cache expired. Re-upload files.")
    frames = dict(entry["frames"])
    workflow = entry["workflow"]

    # Apply column-remap so workflow steps that reference original column names still work.
    # Map shape: { slot: { actual_col_name_in_today_file → original_name_workflow_expects } }
    for slot, mapping in (column_remap or {}).items():
        if slot in frames and isinstance(mapping, dict):
            frames[slot] = frames[slot].rename(columns=mapping)

    result = _execute(workflow, frames, up_to=step_index)
    return {
        "run_id": run_id,
        "step_results": result["step_results"],
    }


@router.post("/run-finalize")
async def run_finalize(body: Dict[str, Any]):
    """Phase 2 final: run the full pipeline and return the resulting xlsx as base64."""
    run_id = body.get("run_id", "")
    column_remap = body.get("column_remap") or {}

    entry = _cache_get(run_id)
    if not entry:
        raise HTTPException(410, "Run cache expired. Re-upload files.")
    frames = dict(entry["frames"])
    workflow = entry["workflow"]

    for slot, mapping in (column_remap or {}).items():
        if slot in frames and isinstance(mapping, dict):
            frames[slot] = frames[slot].rename(columns=mapping)

    result = _execute(workflow, frames, up_to=None)
    output = workflow.get("output", {}) or {}
    final_step = output.get("final_step")
    if not final_step or final_step not in result["frames"]:
        raise HTTPException(400, f"Final step '{final_step}' not found in workflow output.")

    df = result["frames"][final_step]

    # Optional column projection
    cols = output.get("columns")
    if cols:
        present = [c for c in cols if c in df.columns]
        df = df[present].copy()

    output_type = (output.get("output_type") or "records").lower()
    sheet = output.get("sheet", "Result") or "Result"
    filename = output.get("filename", "result.xlsx")

    # Value mode — single-cell aggregation (KPI). Returns scalar + a one-cell xlsx.
    if output_type == "value":
        agg_func = (output.get("agg_func") or "count").lower()
        agg_column = output.get("agg_column")
        scalar_value: Any
        try:
            if agg_func == "count" and not agg_column:
                scalar_value = int(len(df))
            else:
                if not agg_column or agg_column not in df.columns:
                    raise HTTPException(400, f"agg_column '{agg_column}' missing for value output")
                series = pd.to_numeric(df[agg_column], errors="coerce")
                if agg_func == "sum":
                    scalar_value = float(series.sum())
                elif agg_func == "avg":
                    scalar_value = float(series.mean())
                elif agg_func == "min":
                    scalar_value = float(series.min())
                elif agg_func == "max":
                    scalar_value = float(series.max())
                elif agg_func == "count":
                    scalar_value = int(series.notna().sum())
                else:
                    raise HTTPException(400, f"Unsupported agg_func '{agg_func}'")
        except HTTPException:
            raise
        except Exception as e:
            raise HTTPException(500, f"Aggregation failed: {e}")

        # Cast to plain int when possible for clean rendering
        if isinstance(scalar_value, float) and scalar_value.is_integer():
            scalar_value = int(scalar_value)

        buf = io.BytesIO()
        try:
            with pd.ExcelWriter(buf, engine="openpyxl") as writer:
                pd.DataFrame([{f"{agg_func.upper()}({agg_column or 'rows'})": scalar_value}]).to_excel(
                    writer, index=False, sheet_name=sheet,
                )
        except Exception as e:
            raise HTTPException(500, f"Excel write failed: {e}")
        return {
            "filename": filename,
            "sheet": sheet,
            "output_type": "value",
            "scalar_value": scalar_value,
            "agg_func": agg_func,
            "agg_column": agg_column,
            "row_count": int(len(df)),
            "columns": [str(c) for c in df.columns],
            "preview_rows": _df_records_safe(df.head(PREVIEW_ROWS)),
            "xlsx_base64": base64.b64encode(buf.getvalue()).decode("ascii"),
        }

    # Template mode — branded title row + frozen header + auto-width columns
    buf = io.BytesIO()
    try:
        with pd.ExcelWriter(buf, engine="openpyxl") as writer:
            if output_type == "template":
                title = output.get("template_title") or filename.rsplit(".", 1)[0]
                # Reserve top row for the title; data starts on row 3 (1-indexed: 1=title, 2=blank, 3=header)
                df.to_excel(writer, index=False, sheet_name=sheet, startrow=2)
                ws = writer.sheets[sheet]
                from openpyxl.styles import Font, PatternFill, Alignment
                last_col_letter = ws.cell(row=3, column=max(1, df.shape[1])).column_letter
                ws.cell(row=1, column=1, value=str(title)[:120])
                ws.cell(row=1, column=1).font = Font(bold=True, size=14, color="FFFFFFFF")
                ws.cell(row=1, column=1).alignment = Alignment(vertical="center")
                ws.cell(row=1, column=1).fill = PatternFill("solid", fgColor="FF7C3AED")
                ws.merge_cells(f"A1:{last_col_letter}1")
                ws.row_dimensions[1].height = 26
                # Header row formatting
                for col_idx in range(1, df.shape[1] + 1):
                    cell = ws.cell(row=3, column=col_idx)
                    cell.font = Font(bold=True, color="FFFFFFFF")
                    cell.fill = PatternFill("solid", fgColor="FF1F2937")
                ws.freeze_panes = "A4"
                # Auto column width (capped) — sample first 200 rows for speed
                for col_idx, col_name in enumerate(df.columns, start=1):
                    sample = df[col_name].astype(str).head(200)
                    max_len = max([len(str(col_name))] + [len(v) for v in sample]) if len(sample) else len(str(col_name))
                    ws.column_dimensions[ws.cell(row=3, column=col_idx).column_letter].width = min(max(10, max_len + 2), 48)
            else:
                df.to_excel(writer, index=False, sheet_name=sheet)
    except Exception as e:
        raise HTTPException(500, f"Excel write failed: {e}")

    return {
        "filename": filename,
        "sheet": sheet,
        "output_type": output_type,
        "row_count": int(len(df)),
        "columns": [str(c) for c in df.columns],
        "preview_rows": _df_records_safe(df.head(PREVIEW_ROWS)),
        "xlsx_base64": base64.b64encode(buf.getvalue()).decode("ascii"),
    }
